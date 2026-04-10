"""Excel report generator using openpyxl."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lbo_simulator.models.schemas import LBOConfigSchema, LBOResultsSchema


class ExcelExporter:
    """Exports LBO results to formatted Excel workbook."""

    # Styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=14, color="2F5496")
    SUBTITLE_FONT = Font(bold=True, size=12, color="2F5496")
    NUMBER_FMT = "#,##0"
    PCT_FMT = "0.0%"
    MULTIPLE_FMT = "0.00x"

    def __init__(self, config: LBOConfigSchema, results: LBOResultsSchema) -> None:
        self.config = config
        self.results = results
        self.wb = openpyxl.Workbook()

    def export(self, output_path: str | Path) -> Path:
        """Create and save Excel workbook.

        Args:
            output_path: Output file path.

        Returns:
            Path to saved file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self._create_summary_sheet()
        self._create_cash_flow_sheet()
        self._create_debt_schedule_sheet()
        self._create_covenant_sheet()

        self.wb.save(str(output_path))
        return output_path

    def _create_summary_sheet(self) -> None:
        ws = self.wb.active
        ws.title = "Summary"

        # Title
        ws["A1"] = "LBO Simulation Summary"
        ws["A1"].font = self.TITLE_FONT

        ws["A2"] = f"Company: {self.config.company.name}"
        ws["A2"].font = Font(italic=True, size=10)

        # Key metrics
        metrics = [
            ("Equity IRR", self.results.irr, self.PCT_FMT),
            ("MOIC", self.results.moic, self.MULTIPLE_FMT),
            ("Payback Period", self.results.payback_period_years, "0.0 years"),
            ("Total Equity Invested", self.results.total_equity_invested, self.NUMBER_FMT),
            ("Total Equity Returned", self.results.total_equity_returned, self.NUMBER_FMT),
            ("Exit Enterprise Value", self.results.exit_enterprise_value, self.NUMBER_FMT),
            ("Exit Equity Value", self.results.exit_equity_value, self.NUMBER_FMT),
            ("Remaining Debt at Exit", self.results.remaining_debt_at_exit, self.NUMBER_FMT),
            ("Total Interest Paid", self.results.total_interest_paid, self.NUMBER_FMT),
            ("Total Principal Repaid", self.results.total_principal_repaid, self.NUMBER_FMT),
        ]

        row = 4
        for label, value, fmt in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            if isinstance(value, float) and "years" not in fmt:
                ws[f"B{row}"].number_format = fmt
            row += 1

        # Sources & Uses
        row += 1
        ws[f"A{row}"] = "Sources & Uses"
        ws[f"A{row}"].font = self.SUBTITLE_FONT
        row += 1

        sources = self.config.sources_and_uses
        source_items = [
            ("Equity Contribution", sources.equity_contribution),
            ("Senior Debt", sources.senior_debt),
            ("Mezzanine Debt", sources.mezzanine_debt),
            ("High Yield Debt", sources.high_yield_debt),
            ("Total Sources", sources.total_sources),
        ]

        for label, value in source_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = self.NUMBER_FMT
            if "Total" in label:
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"].font = Font(bold=True)
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_cash_flow_sheet(self) -> None:
        ws = self.wb.create_sheet("Cash Flows")

        ws["A1"] = "Annual Cash Flows"
        ws["A1"].font = self.TITLE_FONT

        # Headers
        headers = [
            "Year",
            "Revenue",
            "EBITDA",
            "Taxes",
            "Capex",
            "ΔNWC",
            "Unlevered FCF",
            "Mandatory Amort",
            "Optional Sweep",
            "PIK Accrued",
            "Equity Distribution",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data
        for i, cf in enumerate(self.results.annual_cash_flows):
            row = 4 + i
            ws.cell(row=row, column=1, value=cf.year)
            ws.cell(row=row, column=2, value=cf.revenue).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=3, value=cf.ebitda).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=4, value=cf.taxes).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=5, value=cf.capex).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=6, value=cf.delta_nwc).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=7, value=cf.unlevered_fcf).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=8, value=cf.mandatory_amortization).number_format = (
                self.NUMBER_FMT
            )
            ws.cell(row=row, column=9, value=cf.optional_sweep).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=10, value=cf.pik_accrued).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=11, value=cf.equity_distribution).number_format = (
                self.NUMBER_FMT
            )

        # Column widths
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_debt_schedule_sheet(self) -> None:
        ws = self.wb.create_sheet("Debt Schedule")

        ws["A1"] = "Debt Schedule by Tranche"
        ws["A1"].font = self.TITLE_FONT

        headers = [
            "Year",
            "Tranche",
            "Beg Balance",
            "Interest Paid",
            "Mandatory Amort",
            "Optional Sweep",
            "PIK Accrued",
            "End Balance",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for i, ds in enumerate(self.results.debt_schedule):
            row = 4 + i
            ws.cell(row=row, column=1, value=ds.year)
            ws.cell(row=row, column=2, value=ds.tranche_name)
            ws.cell(row=row, column=3, value=ds.beginning_balance).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=4, value=ds.interest_paid).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=5, value=ds.mandatory_amortization).number_format = (
                self.NUMBER_FMT
            )
            ws.cell(row=row, column=6, value=ds.optional_sweep).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=7, value=ds.pik_accrued).number_format = self.NUMBER_FMT
            ws.cell(row=row, column=8, value=ds.ending_balance).number_format = self.NUMBER_FMT

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_covenant_sheet(self) -> None:
        ws = self.wb.create_sheet("Covenants")

        ws["A1"] = "Covenant Compliance"
        ws["A1"].font = self.TITLE_FONT

        if not self.results.covenant_breaches:
            ws["A3"] = "No covenant breaches detected"
            ws["A3"].font = Font(color="006400", bold=True, size=12)
            return

        headers = ["Year", "Covenant", "Actual", "Threshold", "Severity", "Remediation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL

        # Color map for severity
        severity_colors = {
            "warning": "FFC000",
            "breach": "FF6600",
            "critical": "FF0000",
        }

        for i, breach in enumerate(self.results.covenant_breaches):
            row = 4 + i
            ws.cell(row=row, column=1, value=breach.year)
            ws.cell(row=row, column=2, value=breach.covenant_name)
            ws.cell(row=row, column=3, value=breach.actual_value).number_format = "0.00"
            ws.cell(row=row, column=4, value=breach.threshold_value).number_format = "0.00"
            severity_cell = ws.cell(row=row, column=5, value=breach.severity)
            severity_cell.fill = PatternFill(
                start_color=severity_colors.get(breach.severity, "CCCCCC"),
                end_color=severity_colors.get(breach.severity, "CCCCCC"),
                fill_type="solid",
            )
            ws.cell(row=row, column=6, value=breach.remediation_applied)
